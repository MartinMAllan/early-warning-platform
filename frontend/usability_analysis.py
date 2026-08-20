"""Scores the SUS usability study (spec objective 6, Front-End & UX component)
from the raw Google Forms export and writes output/usability_summary.json.

Run after adding/refreshing responses in
tests/Usability Evaluation Questionnaire - Early Warning Platform (Responses).xlsx:

    python frontend/usability_analysis.py

Standard System Usability Scale (SUS) scoring: 10 statements, 1-5 Likert
scale, alternating positive/negative phrasing. Odd items (1,3,5,7,9)
contribute (score - 1); even items (2,4,6,8,10) contribute (5 - score).
Sum * 2.5 gives a 0-100 SUS score per respondent. A mean above ~68 is
"above average" against the published SUS benchmark (Bangor, Kortum &
Miller, 2008).
"""

import json
import re
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
SURVEY = ROOT / "tests" / "Usability Evaluation Questionnaire – Early Warning Platform (Responses).xlsx"
OUT = ROOT / "output" / "usability_summary.json"

SUS_COL_START, SUS_COL_END = 5, 15  # zero-indexed columns of the 10 SUS items
ROLE_COL = 1


def parse_likert(value):
    """Cell may be a plain number or Google Forms' "N – Label" text."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.match(r"\s*(\d+)", str(value))
    return float(match.group(1)) if match else None


def sus_score(items):
    total = 0.0
    for i, raw in enumerate(items):
        v = parse_likert(raw)
        if v is None:
            return None
        total += (v - 1) if i % 2 == 0 else (5 - v)
    return total * 2.5


def main():
    wb = openpyxl.load_workbook(SURVEY, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    data = rows[1:]

    per_respondent = []
    for row in data:
        score = sus_score(row[SUS_COL_START:SUS_COL_END])
        per_respondent.append({"role": row[ROLE_COL], "sus_score": score})

    valid = [r["sus_score"] for r in per_respondent if r["sus_score"] is not None]
    summary = {
        "source": SURVEY.name,
        "n_responses": len(data),
        "n_scored": len(valid),
        "mean_sus_score": round(sum(valid) / len(valid), 1) if valid else None,
        "min_sus_score": min(valid) if valid else None,
        "max_sus_score": max(valid) if valid else None,
        "benchmark_note": "Mean SUS >= 68 is 'above average' per Bangor, Kortum & Miller (2008)",
        "roles_represented": sorted({r["role"] for r in per_respondent if r["role"]}),
        "per_respondent": per_respondent,
    }
    OUT.write_text(json.dumps(summary, indent=2))
    print(f"Scored {len(valid)}/{len(data)} responses, mean SUS = {summary['mean_sus_score']}")
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
